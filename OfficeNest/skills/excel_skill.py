"""
Excel Skill：8 个 LLM 可调用的 Excel 工具。
LLM 传参可能不准确（file_path/filepath、sheet_name/sheet），自动归一化。
"""
from pathlib import Path
from skills.base import BaseSkill
from adapters.office_bridge import OfficeBridge
from core.interfaces.office_bridge import CellRange


class ExcelSkill(BaseSkill):

    # LLM 常见的参数名变体
    _ALIASES = {
        "filepath": ["filepath", "file_path", "path", "filename"],
        "sheet": ["sheet", "sheet_name", "sheetname"],
        "range": ["range", "cell_range", "region"],
        "start_cell": ["start_cell", "start", "cell"],
        "headers": ["headers", "header", "columns", "cols"],
        "data": ["data", "values", "rows", "content"],
        "rows": ["rows", "data", "values"],
        "column": ["column", "col", "col_name"],
        "value": ["value", "val", "search", "keyword"],
        "sheet_name": ["sheet_name", "sheetname", "name", "sheet"],
    }

    @classmethod
    def _norm(cls, kwargs: dict) -> dict:
        """归一化参数名：把 LLM 的各种变体统一为标准名。"""
        result = {}
        for std_name, aliases in cls._ALIASES.items():
            for alias in aliases:
                if alias in kwargs:
                    result[std_name] = kwargs[alias]
                    break
        # 保留未识别的参数
        for k, v in kwargs.items():
            found = False
            for aliases in cls._ALIASES.values():
                if k in aliases:
                    found = True
                    break
            if not found:
                result[k] = v
        return result

    def __init__(self, office: OfficeBridge, registry=None, workbooks_dir: str = "./workbooks"):
        self._office = office
        self._registry = registry  # FileRegistry, 可选
        self._locked_file = ""     # 锁定文件路径，由面板设置
        self._workbooks_dir = workbooks_dir  # 从配置读取，不再硬编码

    def set_locked_file(self, path: str):
        """设置文件锁——锁定后所有操作只能针对此文件。"""
        self._locked_file = path.strip('"').strip("'").strip() if path else ""

    def _check_lock(self, filepath: str) -> str:
        """检查文件锁。返回错误消息或空字符串（通过）。"""
        if not self._locked_file:
            return ""
        from pathlib import Path
        locked = Path(self._locked_file).resolve()
        # 清理传入路径的引号
        clean_fp = filepath.strip('"').strip("'").strip() if filepath else ""
        target = Path(clean_fp) if Path(clean_fp).is_absolute() else Path(self._workbooks_dir) / clean_fp
        try:
            target = target.resolve()
        except Exception:
            pass
        # 比较文件名（resolve 后路径可能有细微差异，用 name+size 兜底）
        if target == locked or target.name == locked.name:
            return ""
        return f"🔒 当前会话已锁定文件: {locked.name}，禁止操作其他文件。请操作锁定文件或解除锁定。"

    @property
    def name(self) -> str:
        return "excel"

    def get_tools(self) -> list[dict]:
        return [
            {"name": "excel_list_files", "fn": self.excel_list_files,
             "schema": self._schema("列出所有已注册的 Excel 文件", {})},
            {"name": "excel_create", "fn": self.excel_create,
             "schema": self._schema("创建新的 Excel 文件", {
                 "filepath": {"type": "string", "description": "文件路径，如 ./workbooks/测试.xlsx"},
                 "headers": {"type": "array", "items": {"type": "string"}, "description": "表头行，如 ['姓名','岗位','评分']"},
             }, ["filepath"])},
            {"name": "excel_read", "fn": self.excel_read,
             "schema": self._schema("读取 Excel 文件的指定区域", {
                 "filepath": {"type": "string", "description": "文件路径"},
                 "sheet": {"type": "string", "description": "Sheet 名称，默认第一个"},
                 "range": {"type": "string", "description": "区域，如 A1:D20，默认 A1 开始的全部区域"},
             }, ["filepath"])},
            {"name": "excel_write", "fn": self.excel_write,
             "schema": self._schema("往 Excel 写入数据。写入前自动检查目标区域，已有数据时需设 overwrite=true 确认覆盖。返回写入验证数据。", {
                 "filepath": {"type": "string", "description": "文件路径"},
                 "sheet": {"type": "string", "description": "Sheet 名称，默认第一个"},
                 "start_cell": {"type": "string", "description": "起始单元格，如 A1"},
                 "data": {"type": "array", "description": "二维数组数据"},
                 "overwrite": {"type": "boolean", "description": "目标区域已有数据时是否覆盖，默认 false"},
             }, ["filepath", "data"])},
            {"name": "excel_add_sheet", "fn": self.excel_add_sheet,
             "schema": self._schema("在现有 Excel 中添加新 Sheet", {
                 "filepath": {"type": "string", "description": "文件路径"},
                 "sheet_name": {"type": "string", "description": "新 Sheet 名称"},
             }, ["filepath", "sheet_name"])},
            {"name": "excel_append_rows", "fn": self.excel_append_rows,
             "schema": self._schema("往 Excel 末尾追加数据行", {
                 "filepath": {"type": "string", "description": "文件路径"},
                 "sheet": {"type": "string", "description": "Sheet 名称，默认第一个"},
                 "rows": {"type": "array", "description": "要追加的行数据"},
             }, ["filepath", "rows"])},
            {"name": "excel_find_row", "fn": self.excel_find_row,
             "schema": self._schema("在指定列中查找匹配的行", {
                 "filepath": {"type": "string", "description": "文件路径"},
                 "sheet": {"type": "string", "description": "Sheet 名称"},
                 "column": {"type": "string", "description": "搜索列，如 A"},
                 "value": {"type": "string", "description": "要查找的值"},
             }, ["filepath", "column", "value"])},
            {"name": "excel_delete_rows", "fn": self.excel_delete_rows,
             "schema": self._schema("删除 Excel 中匹配的整行。危险操作，会永久删除数据。", {
                 "filepath": {"type": "string", "description": "文件路径"},
                 "sheet": {"type": "string", "description": "Sheet 名称"},
                 "column": {"type": "string", "description": "匹配列，如 A"},
                 "value": {"type": "string", "description": "匹配值，删除该列等于此值的所有行"},
             }, ["filepath", "column", "value"])},
            {"name": "excel_get_sheets", "fn": self.excel_get_sheets,
             "schema": self._schema("获取 Excel 的所有 Sheet 名称", {
                 "filepath": {"type": "string", "description": "文件路径"},
             }, ["filepath"])},
            # ── 格式化工具 ──
            {"name": "excel_format_range", "fn": self.excel_format_range,
             "schema": self._schema("设置单元格格式：字体/对齐/边框/列宽/行高/合并/筛选/冻结", {
                 "filepath": {"type": "string", "description": "文件路径"},
                 "sheet": {"type": "string", "description": "Sheet 名称"},
                 "range": {"type": "string", "description": "区域，如 A1:D20"},
                 "bold": {"type": "boolean", "description": "是否加粗"},
                 "font_size": {"type": "integer", "description": "字体大小如 12"},
                 "font_color": {"type": "string", "description": "字体颜色如 FF0000"},
                 "bg_color": {"type": "string", "description": "背景色如 FFFF00"},
                 "merge": {"type": "boolean", "description": "是否合并单元格"},
                 "border_style": {"type": "string", "description": "边框: thin/medium/thick/none"},
                 "col_width": {"type": "number", "description": "列宽"},
                 "row_height": {"type": "number", "description": "行高"},
                 "number_format": {"type": "string", "description": "数字格式如 #,##0.00"},
                 "alignment": {"type": "string", "description": "对齐: left/center/right"},
                 "auto_filter": {"type": "boolean", "description": "是否启用自动筛选"},
                 "freeze_panes": {"type": "string", "description": "冻结窗格位置如 A2"},
             }, ["filepath", "range"])},
        ]

    # ── 工具实现 ──

    def _fix_path(self, filepath: str) -> Path:
        """智能路径修正：已有文件直接使用，新文件默认放 workbooks_dir/。"""
        path = Path(filepath.strip('"').strip("'").strip() if filepath else "")
        # 如果路径已存在（文件或绝对路径），直接用
        if path.exists():
            return path.resolve()
        # 尝试 workbooks_dir/ 下找
        wb_path = Path(self._workbooks_dir) / path.name
        if wb_path.exists():
            return wb_path.resolve()
        # 全盘搜索（Everything CLI 优先，glob 兜底）
        searched = self._search_file(path.name)
        if searched:
            return Path(searched).resolve()
        # 空路径/目录 → 自动生成
        if str(path) in ('.', './', '') or path.is_dir():
            import datetime
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            return Path(f"{self._workbooks_dir}/新建表格_{ts}.xlsx").resolve()
        # 无后缀 → 补 .xlsx
        if path.suffix not in ('.xlsx', '.xls', '.xlsm'):
            path = path.with_suffix('.xlsx')
        # 相对路径 → 默认放 workbooks_dir/
        if not path.is_absolute():
            path = Path(self._workbooks_dir) / path.name
        return path.resolve()

    @staticmethod
    def _search_file(filename: str) -> str:
        """全盘搜索文件：优先 Everything CLI，兜底 glob。"""
        if not filename:
            return ""
        # Everything CLI
        try:
            import subprocess, shutil
            if shutil.which("es"):
                result = subprocess.run(
                    ["es", filename], capture_output=True, text=True, timeout=3
                )
                if result.stdout.strip():
                    first = result.stdout.strip().split("\n")[0].strip()
                    if Path(first).exists():
                        return first
        except Exception:
            pass
        # glob 兜底：常见目录
        try:
            import glob
            for pattern in [
                f"C:/Users/*/Downloads/{filename}",
                f"C:/Users/*/Desktop/{filename}",
                f"C:/Users/*/Documents/{filename}",
                f"./**/{filename}",
            ]:
                matches = glob.glob(pattern, recursive=True)
                if matches:
                    return matches[0]
        except Exception:
            pass
        return ""

    async def excel_list_files(self) -> str:
        from adapters.file_bridge import FileBridge
        fb = FileBridge()
        files = fb.list_dir(self._workbooks_dir, "*.xlsx") + fb.list_dir(self._workbooks_dir, "*.xls")
        if not files:
            return "（workbooks/ 目录下暂无 Excel 文件）"
        return "\n".join(f"  {f.name}" for f in files)

    async def excel_create(self, filepath: str = "", headers: list[str] | None = None, **kwargs) -> str:
        kw = self._norm({**kwargs, "filepath": filepath or kwargs.get("file_path", ""), "headers": headers})
        filepath = kw.get("filepath", "")
        lock_err = self._check_lock(filepath)
        if lock_err:
            return lock_err
        headers = kw.get("headers", [])
        path = self._fix_path(filepath)
        path.parent.mkdir(parents=True, exist_ok=True)
        sid = await self._office.excel_open(str(path))
        if headers:
            sheet_data = [[h for h in headers]]
            await self._office.excel_write(sid, CellRange("Sheet", "A1"), sheet_data)
        await self._office.excel_save_as(sid, str(path))
        await self._office.excel_close(sid)

        # 自动注册到文件注册表
        desc = kwargs.get("description", "")
        if not desc and headers:
            desc = f"含 {len(headers)} 列: {', '.join(str(h) for h in headers[:5])}"
        if self._registry:
            from core.interfaces.file_registry import FileEntry
            self._registry.register(FileEntry(
                name=path.name,
                path=str(path.resolve()),
                file_type="excel",
                metadata={"description": desc},
            ))

        return f"✅ Excel 已创建: {path.name}" + (f"，含 {len(headers)} 列表头" if headers else "")

    async def excel_read(self, filepath: str = "", sheet: str = "", range: str = "", **kwargs) -> str:
        kw = self._norm({**kwargs, "filepath": filepath, "sheet": sheet, "range": range})
        filepath = kw.get("filepath", "")
        lock_err = self._check_lock(filepath)
        if lock_err:
            return lock_err
        sheet = kw.get("sheet", "")
        range = kw.get("range", "")
        path = self._fix_path(filepath)

        sid = await self._office.excel_open(str(path))
        try:
            from adapters.file_bridge import FileBridge
            fb = FileBridge()
            if not fb.exists(str(path)):
                return f"❌ 文件不存在: {path.name}"

            # 先获取 Sheet 列表
            if not sheet:
                import openpyxl
                wb = openpyxl.load_workbook(str(path), read_only=True)
                sheet = wb.sheetnames[0]
                wb.close()

            end = range.split(":")[1] if ":" in range else ""
            # 无 range 时自动读全表（非 A1）
            if not range:
                try:
                    wb2 = openpyxl.load_workbook(str(path), read_only=True)
                    ws2 = wb2[sheet]
                    max_col = ws2.max_column
                    max_row = ws2.max_row
                    wb2.close()
                    col_letter = self._office._col_letter(max_col)
                    end = f"{col_letter}{max_row}"
                except Exception:
                    end = ""
            cell_range = CellRange(sheet, range.split(":")[0] if range else "A1", end)
            data = await self._office.excel_read(sid, cell_range)

            if not data:
                return "（空区域）"

            # 提取表头（第一行）作为列映射
            header_line = ""
            if data:
                from adapters.office_bridge import OfficeBridge as _OB
                hdr = data[0]
                col_letters = []
                for i, v in enumerate(hdr):
                    letter = self._office._col_letter(i + 1)
                    col_letters.append(f"{letter}={v}" if v is not None else f"{letter}=(空)")
                header_line = "表头: " + " | ".join(col_letters[:10])

            # 格式化输出（最多50行）
            lines = []
            for i, row in enumerate(data[:50]):
                lines.append(" | ".join(str(v) if v is not None else "" for v in row))
            result = header_line + "\n" + "\n".join(lines)
            if len(data) > 50:
                result += f"\n... （共 {len(data)} 行，仅显示前 50 行）"
            return result
        finally:
            await self._office.excel_close(sid)

    async def excel_write(self, filepath: str = "", data: list[list] | None = None, sheet: str = "", start_cell: str = "A1", **kwargs) -> str:
        kw = self._norm({**kwargs, "filepath": filepath, "data": data, "sheet": sheet, "start_cell": start_cell})
        filepath = kw.get("filepath", "")
        lock_err = self._check_lock(filepath)
        if lock_err:
            return lock_err
        data = kw.get("data", [])
        sheet = kw.get("sheet", "")
        start_cell = kw.get("start_cell", "A1")
        overwrite = kw.get("overwrite", False)
        path = self._fix_path(filepath)
        sid = await self._office.excel_open(str(path))
        try:
            # 不传 sheet 时取实际第一个 Sheet 名（不默认 "Sheet"）
            if not sheet:
                import openpyxl
                wb_tmp = openpyxl.load_workbook(str(path), read_only=True)
                sheet = wb_tmp.sheetnames[0]
                wb_tmp.close()

            # 写入前检查目标行是否已有数据
            if not overwrite:
                start_col, start_row = self._office._parse_cell(start_cell)
                n_rows = len(data) if data else 0
                check_end = f"{self._office._col_letter(start_col + max(len(data[0]) if data else 1, 1) - 1)}{start_row + n_rows - 1}" if n_rows else start_cell
                existing = await self._office.excel_read(sid, CellRange(sheet, start_cell, check_end))
                has_data = any(any(str(v).strip() for v in row if v is not None) for row in existing)
                if has_data:
                    preview = " | ".join(str(v) if v is not None else "" for v in existing[0][:6])
                    return f"⚠️ 目标区域已有数据: [{preview}]\n确认覆盖请设 overwrite=true"

            cell_range = CellRange(sheet, start_cell)
            rows = await self._office.excel_write(sid, cell_range, data)
            await self._office.excel_save(sid)

            # 写后回读验证
            verify = await self._office.excel_read(sid, CellRange(sheet, start_cell))
            verify_line = " | ".join(str(v) if v is not None else "" for v in verify[0][:8]) if verify else ""

            # 读表头（第1行）
            headers_line = ""
            try:
                hdr = await self._office.excel_read(sid, CellRange(sheet, "A1"))
                if hdr:
                    headers_line = "表头: " + " | ".join(str(v) if v is not None else "" for v in hdr[0][:8])
            except Exception:
                pass

            result = f"✅ 已写入 {rows} 行到 {sheet}!{start_cell}"
            if headers_line:
                result += f"\n   {headers_line}"
            if verify_line:
                result += f"\n   写入验证: {verify_line}"
            return result
        finally:
            await self._office.excel_close(sid)

    async def excel_add_sheet(self, filepath: str = "", sheet_name: str = "", **kwargs) -> str:
        kw = self._norm({**kwargs, "filepath": filepath, "sheet_name": sheet_name})
        filepath = kw.get("filepath", "")
        lock_err = self._check_lock(filepath)
        if lock_err:
            return lock_err
        sheet_name = kw.get("sheet_name", "")
        path = self._fix_path(filepath)
        sid = await self._office.excel_open(str(path))
        try:
            name = await self._office.excel_add_sheet(sid, sheet_name)
            await self._office.excel_save(sid)
            return f"✅ 已添加 Sheet: {name}"
        finally:
            await self._office.excel_close(sid)

    async def excel_append_rows(self, filepath: str = "", rows: list[list] | None = None, sheet: str = "", **kwargs) -> str:
        kw = self._norm({**kwargs, "filepath": filepath, "rows": rows, "sheet": sheet})
        filepath = kw.get("filepath", "")
        lock_err = self._check_lock(filepath)
        if lock_err:
            return lock_err
        rows = kw.get("rows", [])
        sheet = kw.get("sheet", "")
        if not rows:
            return "❌ 没有要追加的数据"
        path = self._fix_path(filepath)
        sid = await self._office.excel_open(str(path))
        try:
            # 不传 sheet 时取实际第一个
            if not sheet:
                import openpyxl
                wb_tmp = openpyxl.load_workbook(str(path), read_only=True)
                sheet = wb_tmp.sheetnames[0]
                wb_tmp.close()

            # 读全表找最后一行
            existing = await self._office.excel_read(sid, CellRange(sheet, "A1"))
            next_row = len(existing) + 1 if existing else 1
            cell_range = CellRange(sheet, f"A{next_row}")
            written = await self._office.excel_write(sid, cell_range, rows)
            await self._office.excel_save(sid)

            # 写后验证
            verify = await self._office.excel_read(sid, CellRange(sheet, f"A{next_row}"))
            verify_data = " | ".join(str(v) if v is not None else "" for v in (verify[0] if verify else [])[:6])
            return f"✅ 已追加 {written} 行（第 {next_row} 行起）| 验证: {verify_data}"
        finally:
            await self._office.excel_close(sid)

    async def excel_find_row(self, filepath: str = "", column: str = "", value: str = "", sheet: str = "", **kwargs) -> str:
        kw = self._norm({**kwargs, "filepath": filepath, "column": column, "value": value, "sheet": sheet})
        filepath = kw.get("filepath", "")
        lock_err = self._check_lock(filepath)
        if lock_err:
            return lock_err
        column = kw.get("column", "")
        value = kw.get("value", "")
        sheet = kw.get("sheet", "")
        path = self._fix_path(filepath)
        sid = await self._office.excel_open(str(path))
        try:
            if not sheet:
                import openpyxl
                wb = openpyxl.load_workbook(str(path), read_only=True)
                sheet = wb.sheetnames[0]
                wb.close()

            cell_range = CellRange(sheet, "A1")
            data = await self._office.excel_read(sid, cell_range)
            if not data:
                return "（文件为空）"

            # 解析列索引
            col_idx = 0
            for ch in column.upper():
                col_idx = col_idx * 26 + (ord(ch) - ord('A'))

            matches = []
            value_str = str(value).strip()
            value_lower = value_str.lower()
            for i, row in enumerate(data):
                if col_idx < len(row):
                    cell_val = str(row[col_idx]).strip() if row[col_idx] is not None else ""
                    # 精确匹配 or 忽略大小写 or 包含匹配
                    if cell_val == value_str or cell_val.lower() == value_lower or value_str in cell_val:
                        matches.append(f"  第{i + 1}行: " + " | ".join(str(v) for v in row))

            if matches:
                return f"找到 {len(matches)} 条匹配:\n" + "\n".join(matches[:10])
            return f"在列 {column} 中未找到 '{value}'。请确认：①名字写法是否正确（空格/繁简）②是否在其他列 ③提供行号"
        finally:
            await self._office.excel_close(sid)

    async def excel_delete_rows(self, filepath: str = "", column: str = "", value: str = "",
                                  sheet: str = "", **kwargs) -> str:
        kw = self._norm({**kwargs, "filepath": filepath, "column": column, "value": value, "sheet": sheet})
        filepath = kw.get("filepath", "")
        lock_err = self._check_lock(filepath)
        if lock_err:
            return lock_err
        column = kw.get("column", "")
        value = kw.get("value", "")
        sheet = kw.get("sheet", "")
        path = self._fix_path(filepath)
        sid = await self._office.excel_open(str(path))
        try:
            if not sheet:
                import openpyxl
                wb = openpyxl.load_workbook(str(path), read_only=True)
                sheet = wb.sheetnames[0]
                wb.close()

            # 读全部数据
            data = await self._office.excel_read(sid, CellRange(sheet, "A1"))
            if not data:
                return "（文件为空，无需删除）"

            # 解析列索引
            col_idx = 0
            for ch in column.upper():
                col_idx = col_idx * 26 + (ord(ch) - ord('A'))

            # 过滤掉匹配行
            deleted = 0
            kept = [data[0]]  # 保留表头
            for row in data[1:]:
                if col_idx < len(row) and str(row[col_idx]) == value:
                    deleted += 1
                else:
                    kept.append(row)

            # 写回（先清空再写）
            empty = [[""] * len(data[0]) for _ in range(len(data))]
            await self._office.excel_write(sid, CellRange(sheet, "A1"), empty)
            await self._office.excel_write(sid, CellRange(sheet, "A1"), kept)
            await self._office.excel_save(sid)

            return f"✅ 已删除 {deleted} 行（列 {column} = '{value}'），剩余 {len(kept) - 1} 行数据"
        finally:
            await self._office.excel_close(sid)

    async def excel_format_range(self, filepath: str = "", range: str = "", **kwargs) -> str:
        """批量设置单元格格式。热模式用 win32com，冷模式用 openpyxl。"""
        kw = self._norm({**kwargs, "filepath": filepath, "range": range})
        filepath = kw.get("filepath", "")
        lock_err = self._check_lock(filepath)
        if lock_err:
            return lock_err
        range_spec = kw.get("range", "A1")
        sheet = kw.get("sheet", "")

        path = Path(filepath.strip('"').strip("'").strip() if filepath else "")
        if not path.is_absolute() and not path.exists():
            path = Path(self._workbooks_dir) / path.name

        # 检测热模式：如果文件被 Excel 锁定，用 win32com
        is_locked = False
        if path.exists():
            try:
                f = open(str(path), "a"); f.close()
            except (IOError, PermissionError):
                is_locked = True

        if is_locked:
            return self._format_hot(str(path), range_spec, sheet, **kwargs)
        else:
            # 清理 kwargs 中的路径参数，避免传入 _format_cold 时重复
            clean_kw = {k: v for k, v in kwargs.items()
                        if k not in ('sheet', 'filepath', 'range', 'sheet_name',
                                     'file_path', 'col', 'column', 'value',
                                     'start_cell', 'data', 'rows', 'headers')}
            return self._format_cold(str(path), range_spec, sheet, **clean_kw)

    def _format_cold(self, path: str, range_spec: str, sheet: str, **kwargs) -> str:

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter

        kwargs.pop("filepath", None)
        kwargs.pop("range", None)
        kwargs.pop("sheet", None)
        kwargs.pop("sheet_name", None)
        kwargs.pop("file_path", None)
        kwargs.pop("col", None)
        kwargs.pop("column", None)
        kwargs.pop("value", None)
        kwargs.pop("start_cell", None)
        kwargs.pop("data", None)
        kwargs.pop("rows", None)
        kwargs.pop("headers", None)

        wb = openpyxl.load_workbook(str(path))
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
        if not sheet:
            sheet = ws.title

        # 解析区域
        parts = range_spec.split(":")
        start_cell = self._office._parse_cell(parts[0])
        end_cell = self._office._parse_cell(parts[1]) if len(parts) > 1 else start_cell
        cells = ws.iter_rows(
            min_row=start_cell[1], max_row=end_cell[1],
            min_col=start_cell[0], max_col=end_cell[0],
        )

        # 构建样式
        font_kw = {}
        if kwargs.get("bold") is not None:
            font_kw["bold"] = bool(kwargs["bold"])
        if kwargs.get("font_size"):
            font_kw["size"] = int(kwargs["font_size"])
        if kwargs.get("font_color"):
            font_kw["color"] = kwargs["font_color"].lstrip("#")

        fill = None
        if kwargs.get("bg_color"):
            fill = PatternFill(start_color=kwargs["bg_color"].lstrip("#"),
                               end_color=kwargs["bg_color"].lstrip("#"), fill_type="solid")

        align = None
        if kwargs.get("alignment"):
            al_map = {"left": "left", "center": "center", "right": "right"}
            align = Alignment(horizontal=al_map.get(kwargs["alignment"], "center"),
                              vertical="center")

        border = None
        if kwargs.get("border_style"):
            bs = kwargs["border_style"]
            side_map = {"thin": "thin", "medium": "medium", "thick": "thick", "none": None}
            side = Side(style=side_map.get(bs, "thin"))
            border = Border(left=side, right=side, top=side, bottom=side)

        changes = []

        for row_cells in cells:
            for cell in row_cells:
                if font_kw:
                    cell.font = Font(**font_kw)
                if fill:
                    cell.fill = fill
                if align:
                    cell.alignment = align
                if border:
                    cell.border = border
                if kwargs.get("number_format"):
                    cell.number_format = str(kwargs["number_format"])

        # 合并单元格
        if kwargs.get("merge") and len(parts) > 1:
            ws.merge_cells(range_spec)
            changes.append("合并")

        # 列宽
        cw = kwargs.get("col_width")
        if cw and str(cw).strip():
            try:
                for c in range(start_cell[0], end_cell[0] + 1):
                    ws.column_dimensions[get_column_letter(c)].width = float(cw)
                changes.append("列宽")
            except (ValueError, TypeError):
                pass

        # 行高
        rh = kwargs.get("row_height")
        if rh and str(rh).strip():
            try:
                for r in range(start_cell[1], end_cell[1] + 1):
                    ws.row_dimensions[r].height = float(rh)
                changes.append("行高")
            except (ValueError, TypeError):
                pass

        # 自动筛选
        if kwargs.get("auto_filter"):
            ws.auto_filter.ref = range_spec
            changes.append("筛选")

        # 冻结窗格
        if kwargs.get("freeze_panes"):
            ws.freeze_panes = kwargs["freeze_panes"]
            changes.append("冻结")

        wb.save(str(path))
        wb.close()

        font_desc = []
        if font_kw:
            font_desc.append(
                f"字体: {'加粗' if font_kw.get('bold') else ''}"
                f"{font_kw.get('size','')}pt {font_kw.get('color','')}".replace("  ", " ").strip()
            )
        if fill:
            font_desc.append(f"底色:{kwargs['bg_color']}")
        changes_str = ",".join(font_desc + changes) if (font_desc or changes) else "样式"
        return f"✅ {sheet}!{range_spec} 格式已应用: {changes_str}"

    def _format_hot(self, path: str, range_spec: str, sheet: str, **kwargs) -> str:
        """win32com 热模式格式化——文件正被 Excel 打开时。"""
        try:
            import win32com.client
            excel = win32com.client.GetActiveObject("Excel.Application")
        except Exception:
            return "❌ 无法连接 Excel"
        target = Path(path).resolve()
        wb = None
        for w in excel.Workbooks:
            try:
                if Path(w.FullName).resolve() == target:
                    wb = w; break
            except Exception:
                continue
        if not wb:
            return "❌ 文件未在 Excel 中打开"
        ws = wb.ActiveSheet
        if sheet:
            for s in wb.Worksheets:
                if s.Name == sheet:
                    ws = s; break
        rng = ws.Range(range_spec)
        changes = []
        if kwargs.get("bold") is not None:
            rng.Font.Bold = bool(kwargs["bold"]); changes.append("加粗")
        if kwargs.get("font_size"):
            rng.Font.Size = int(kwargs["font_size"]); changes.append("字号")
        if kwargs.get("font_color"):
            rng.Font.Color = int(kwargs["font_color"].lstrip("#"), 16); changes.append("字体色")
        if kwargs.get("bg_color"):
            rng.Interior.Color = int(kwargs["bg_color"].lstrip("#"), 16); changes.append("底色")
        if kwargs.get("number_format"):
            rng.NumberFormat = str(kwargs["number_format"]); changes.append("数字格式")
        if kwargs.get("merge") and ":" in range_spec:
            rng.Merge(); changes.append("合并")
        if kwargs.get("alignment"):
            al = {"left": -4131, "center": -4108, "right": -4152}
            rng.HorizontalAlignment = al.get(kwargs["alignment"], -4108); changes.append("对齐")
        if kwargs.get("col_width"):
            for c in rng.Columns: c.ColumnWidth = float(kwargs["col_width"])
            changes.append("列宽")
        if kwargs.get("row_height"):
            for r in rng.Rows: r.RowHeight = float(kwargs["row_height"])
            changes.append("行高")
        if kwargs.get("auto_filter"):
            rng.AutoFilter(); changes.append("筛选")
        return f"✅ [win32com] {ws.Name}!{range_spec} 格式实时应用: {', '.join(changes) if changes else '样式'}"

    async def excel_get_sheets(self, filepath: str = "", **kwargs) -> str:
        kw = self._norm({**kwargs, "filepath": filepath})
        filepath = kw.get("filepath", "")
        lock_err = self._check_lock(filepath)
        if lock_err:
            return lock_err
        path = self._fix_path(filepath)
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True)
            sheets = wb.sheetnames
            wb.close()
            return "Sheets: " + ", ".join(sheets)
        except Exception as e:
            return f"❌ 无法读取: {e}"

    # ── helpers ──

    @staticmethod
    def _schema(desc: str, properties: dict, required: list[str] | None = None) -> dict:
        return {
            "type": "function",
            "function": {
                "description": desc,
                "parameters": {
                    "type": "object",
                    "properties": properties,
                    "required": required or [],
                }
            }
        }
