"""
图表 Skill：从指定数据区域生成 Excel 图表。
openpyxl 支持柱状图、折线图、饼图等。
"""
from pathlib import Path
from skills.base import BaseSkill


class ChartSkill(BaseSkill):

    @property
    def name(self) -> str:
        return "chart"

    def get_tools(self) -> list[dict]:
        return [
            {"name": "excel_generate_chart", "fn": self.excel_generate_chart,
             "schema": self._schema("在 Excel 指定数据区域生成图表（柱状图/折线图/饼图）", {
                 "filepath": {"type": "string", "description": "文件路径"},
                 "sheet": {"type": "string", "description": "Sheet 名称"},
                 "data_range": {"type": "string", "description": "数据区域如 A1:B10（第1列=标签，第2列=数值）"},
                 "chart_type": {"type": "string", "description": "图表类型: bar/line/pie"},
                 "title": {"type": "string", "description": "图表标题"},
                 "position": {"type": "string", "description": "图表左上角位置如 E2，默认数据右侧"},
             }, ["filepath", "data_range", "chart_type"])},
        ]

    def excel_generate_chart(self, filepath: str, data_range: str, chart_type: str,
                             title: str = "", sheet: str = "", position: str = "") -> str:
        """生成图表。"""
        path = Path(filepath)
        if not path.is_absolute():
            path = Path("./workbooks") / path.name

        import openpyxl
        from openpyxl.chart import BarChart, LineChart, PieChart, Reference

        wb = openpyxl.load_workbook(str(path))
        ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active

        # 解析数据区域
        parts = data_range.split(":")
        if len(parts) == 2:
            start, end = parts[0], parts[1]
        else:
            return "❌ data_range 格式错误，需要 A1:B10 格式"

        # 创建图表
        chart_map = {
            "bar": BarChart, "line": LineChart, "pie": PieChart,
            "柱状图": BarChart, "折线图": LineChart, "饼图": PieChart,
        }
        chart_cls = chart_map.get(chart_type.lower())
        if not chart_cls:
            return f"❌ 不支持的图表类型: {chart_type}，支持: bar/line/pie"

        chart = chart_cls()
        chart.title = title or chart_type

        # 数据引用（第1列=标签，其余列=数据系列）
        from openpyxl.utils import get_column_letter
        start_coords = self._parse_coords(start)
        end_coords = self._parse_coords(end)
        start_col_letter = get_column_letter(start_coords[0])
        end_col_letter = get_column_letter(end_coords[0])

        # 类别（第一列）
        cats = Reference(ws, min_col=start_coords[0], min_row=start_coords[1] + 1,
                         max_row=end_coords[1])
        chart.set_categories(cats)

        # 数据系列（其余列）
        for col_idx in range(start_coords[0] + 1, end_coords[0] + 1):
            data_ref = Reference(ws, min_col=col_idx, min_row=start_coords[1],
                                 max_row=end_coords[1])
            chart.add_data(data_ref, titles_from_data=True)

        # 位置
        if position:
            pos = position
        else:
            pos = f"{get_column_letter(end_coords[0] + 2)}2"
        ws.add_chart(chart, pos)

        wb.save(str(path))
        wb.close()
        return f"✅ {chart_type}图表已生成在 {path.name} 的 {pos} 位置"

    @staticmethod
    def _parse_coords(cell: str) -> tuple[int, int]:
        """A1 → (col, row)"""
        col_str = ""
        row_str = ""
        for ch in cell.upper():
            if ch.isalpha():
                col_str += ch
            else:
                row_str += ch
        col = 0
        for ch in col_str:
            col = col * 26 + (ord(ch) - ord('A') + 1)
        return col, int(row_str)

    @staticmethod
    def _schema(desc: str, properties: dict, required: list[str] | None = None) -> dict:
        return {"type": "function", "function": {
            "description": desc,
            "parameters": {"type": "object", "properties": properties, "required": required or []}
        }}
