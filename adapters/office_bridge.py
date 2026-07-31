"""
Office Bridge 实现：openpyxl（冷数据）+ win32com（热操作）。
实现 IOfficeBridge 接口。
"""
import os
import time
import uuid
import threading
from pathlib import Path
from typing import Any

from core.interfaces.office_bridge import IOfficeBridge, CellRange, TableData


class OfficeBridge(IOfficeBridge):
    """
    Excel + Word 操作桥接。
    每个 open 返回 session_id，后续操作都带 session_id。
    """

    _com_initialized = threading.local()

    def __init__(self, auto_backup: bool = True):
        self._sessions: dict[str, dict] = {}
        self._auto_backup = auto_backup

    @classmethod
    def _com_dispatch(cls, ms_name: str, wps_name: str):
        """优先 MS Office，失败降级 WPS。"""
        import win32com.client
        try:
            return win32com.client.Dispatch(ms_name)
        except Exception:
            try:
                return win32com.client.Dispatch(wps_name)
            except Exception:
                raise Exception(f"COM 调度失败: {ms_name} / {wps_name}")

    @classmethod
    def _com_get_active(cls, ms_name: str, wps_name: str):
        """优先 MS Office，失败降级 WPS。"""
        import win32com.client
        try:
            return win32com.client.GetActiveObject(ms_name)
        except Exception:
            try:
                return win32com.client.GetActiveObject(wps_name)
            except Exception:
                raise Exception(f"COM 未找到活动对象: {ms_name} / {wps_name}")

    @classmethod
    def _ensure_com(cls):
        """确保当前线程已初始化 COM。"""
        if not getattr(cls._com_initialized, "done", False):
            try:
                import pythoncom
                pythoncom.CoInitialize()
            except ImportError:
                pass
            cls._com_initialized.done = True

    # ═══════════════════════════════════════════════
    # Excel
    # ═══════════════════════════════════════════════

    async def excel_open(self, filepath: str, visible: bool = False) -> str:
        self._ensure_com()
        sid = f"xl_{uuid.uuid4().hex[:8]}"
        path = Path(filepath).resolve()

        # ── 第1步：检测文件是否被 Excel 进程锁定 ──
        locked_by_excel = False
        if path.exists():
            locked_by_excel = self._is_file_locked_by_excel(str(path))

        # ── 第2步：尝试 win32com 热连接已打开的文件 ──
        if path.exists():
            wb, app = self._find_open_workbook(str(path))
            if wb:
                print(f"  🔥 [win32com] 热连接已打开的 Excel: {path.name}")
                self._sessions[sid] = {"type": "excel", "mode": "hot", "wb": wb, "app": app, "path": str(path)}
                return sid

        # ── 第3步：文件被锁定但没找到 → 用 win32com 后台打开 ──
        if locked_by_excel and path.exists():
            try:
                import win32com.client
                app = self._com_dispatch("Excel.Application", "ET.Application")
                app.Visible = visible
                app.DisplayAlerts = False
                wb = app.Workbooks.Open(str(path))
                print(f"  🔥 [win32com] 后台打开锁定的 Excel: {path.name}")
                self._sessions[sid] = {"type": "excel", "mode": "hot", "wb": wb, "app": app, "path": str(path)}
                return sid
            except Exception:
                pass

        # ── 第4步：冷数据模式 openpyxl ──
        if path.exists():
            # 4a: 标准模式
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(path), data_only=True)
                self._sessions[sid] = {"type": "excel", "mode": "cold", "wb": wb, "path": str(path)}
                return sid
            except Exception:
                pass
            # 4b: read_only 模式（更宽松，支持更多格式）
            try:
                import openpyxl
                wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
                self._sessions[sid] = {"type": "excel", "mode": "cold", "wb": wb, "path": str(path)}
                return sid
            except Exception:
                pass
            # 4c: 复制到临时文件再读（绕过文件锁）
            try:
                import tempfile, shutil
                tmp = Path(tempfile.gettempdir()) / f"mother_tmp_{path.name}"
                shutil.copy2(str(path), str(tmp))
                import openpyxl
                wb = openpyxl.load_workbook(str(tmp), data_only=True)
                self._sessions[sid] = {"type": "excel", "mode": "cold", "wb": wb, "path": str(path), "_tmp": str(tmp)}
                return sid
            except Exception:
                pass

        # ── 第5步：新文件 ──
        import openpyxl
        wb = openpyxl.Workbook()
        self._sessions[sid] = {"type": "excel", "mode": "cold", "wb": wb, "path": str(path)}
        return sid

    async def excel_read(self, session_id: str, range_spec: CellRange) -> list[list]:
        s = self._get_session(session_id, "excel")
        wb = s["wb"]

        if s["mode"] == "hot":
            return self._excel_read_hot(s, range_spec)
        else:
            return self._excel_read_cold(wb, range_spec)

    def _excel_read_cold(self, wb, range_spec: CellRange) -> list[list]:
        """openpyxl 读取"""
        ws = wb[range_spec.sheet] if range_spec.sheet in wb.sheetnames else wb.active

        start_col, start_row = self._parse_cell(range_spec.start)
        if range_spec.end:
            end_col, end_row = self._parse_cell(range_spec.end)
        else:
            # 无 end → 自动读全表（从 start 到 used range 边界）
            end_col, end_row = ws.max_column, ws.max_row

        data = []
        for row in ws.iter_rows(min_row=start_row, max_row=end_row,
                                min_col=start_col, max_col=end_col,
                                values_only=True):
            data.append(list(row))
        return data

    def _excel_read_hot(self, s: dict, range_spec: CellRange) -> list[list]:
        """win32com 读取已打开的文件"""
        ws = s["wb"].ActiveSheet
        if range_spec.sheet:
            for sh in s["wb"].Worksheets:
                if sh.Name == range_spec.sheet:
                    ws = sh
                    break

        if range_spec.end:
            addr = f"{range_spec.start}:{range_spec.end}"
            rng = ws.Range(addr)
            if rng.Count == 1:
                return [[rng.Value]]
            return [list(r) for r in rng.Value]
        else:
            # 无 end → 自动读全表（从 start 到 UsedRange 边界）
            used = ws.UsedRange
            start_col, start_row = self._parse_cell(range_spec.start)
            end_col, end_row = used.Columns.Count, used.Rows.Count
            addr = f"{range_spec.start}:{self._col_letter(end_col)}{end_row}"
            rng = ws.Range(addr)
            if rng.Count == 1:
                return [[rng.Value]]
            return [list(r) for r in rng.Value]

    async def excel_write(self, session_id: str, range_spec: CellRange, data: list[list]) -> int:
        s = self._get_session(session_id, "excel")

        if s["mode"] == "hot":
            try:
                return self._excel_write_hot(s, range_spec, data)
            except Exception as e:
                # 热模式写入失败（OLE 错误 / Excel 正在编辑单元格等）→ 切冷模式
                print(f"  ⚠️ [win32com] 写入失败，切冷模式: {e}")
                s["mode"] = "cold"
                # 用 openpyxl 重新打开
                import openpyxl
                s["wb"] = openpyxl.load_workbook(s["path"])
                return self._excel_write_cold(s, range_spec, data)
        else:
            return self._excel_write_cold(s, range_spec, data)

    def _excel_write_cold(self, s: dict, range_spec: CellRange, data: list[list]) -> int:
        wb = s["wb"]
        ws = wb[range_spec.sheet] if range_spec.sheet in wb.sheetnames else wb.active

        start_col, start_row = self._parse_cell(range_spec.start)
        rows_written = 0
        for i, row_data in enumerate(data):
            for j, val in enumerate(row_data):
                ws.cell(row=start_row + i, column=start_col + j, value=val)
            rows_written += 1
        return rows_written

    def _excel_write_hot(self, s: dict, range_spec: CellRange, data: list[list]) -> int:
        ws = s["wb"].ActiveSheet
        if range_spec.sheet:
            for sh in s["wb"].Worksheets:
                if sh.Name == range_spec.sheet:
                    ws = sh
                    break

        start_col, start_row = self._parse_cell(range_spec.start)
        for i, row_data in enumerate(data):
            for j, val in enumerate(row_data):
                col_letter = self._col_letter(start_col + j)
                cell = ws.Range(f"{col_letter}{start_row + i}")
                cell.Value = val
        return len(data)

    async def excel_add_sheet(self, session_id: str, name: str) -> str:
        s = self._get_session(session_id, "excel")
        if s["mode"] == "hot":
            s["wb"].Worksheets.Add().Name = name
        else:
            s["wb"].create_sheet(name)
        return name

    async def excel_save(self, session_id: str) -> None:
        s = self._get_session(session_id, "excel")
        if s["mode"] == "hot":
            s["wb"].Save()
        elif s["mode"] == "cold" and s.get("path"):
            import openpyxl
            s["wb"].save(s["path"])

    async def excel_save_as(self, session_id: str, filepath: str) -> None:
        s = self._get_session(session_id, "excel")
        if s["mode"] == "hot":
            s["wb"].SaveAs(filepath)
        elif s["mode"] == "cold":
            import openpyxl
            s["wb"].save(filepath)

    async def excel_close(self, session_id: str) -> None:
        s = self._get_session(session_id, "excel")
        if s["mode"] == "hot":
            # 热模式不强制关闭用户的 Excel
            pass
        elif s["mode"] == "cold":
            try:
                s["wb"].close()
            except Exception:
                pass
            # 清理临时文件
            tmp = s.get("_tmp")
            if tmp:
                try:
                    Path(tmp).unlink(missing_ok=True)
                except Exception:
                    pass
        self._sessions.pop(session_id, None)

    # ═══════════════════════════════════════════════
    # Word
    # ═══════════════════════════════════════════════

    async def word_open(self, filepath: str, visible: bool = False) -> str:
        self._ensure_com()
        sid = f"wd_{uuid.uuid4().hex[:8]}"
        path = Path(filepath).resolve()

        try:
            import win32com.client

            # 第1步：尝试连接到已在运行的 Word（热模式）
            try:
                word = self._com_get_active("Word.Application", "WPS.Application")
                for d in word.Documents:
                    try:
                        if Path(d.FullName).resolve() == path:
                            print(f"  🔥 [win32com] 热连接 Word: {path.name}")
                            self._sessions[sid] = {"type": "word", "app": word, "doc": d, "path": str(path)}
                            return sid
                    except Exception:
                        continue
            except Exception:
                pass

            # 第2步：启动新的 Word 进程
            word = self._com_dispatch("Word.Application", "WPS.Application")
            word.Visible = False
            word.ScreenUpdating = False
            print(f"  🔥 [win32com] 后台启动 Word 进程")

            if path.exists():
                doc = word.Documents.Open(str(path))
                # .doc → .docx 自动转换
                if path.suffix.lower() == '.doc':
                    docx_path = path.with_suffix('.docx')
                    doc.SaveAs2(str(docx_path), 16)  # 16 = wdFormatDocumentDefault
                    doc.Close()
                    doc = word.Documents.Open(str(docx_path))
                    print(f"  🔥 [win32com] .doc → .docx 已转换: {docx_path.name}")
                    path = docx_path
            else:
                doc = word.Documents.Add()

            self._sessions[sid] = {"type": "word", "app": word, "doc": doc, "path": str(path)}
            return sid
        except ImportError:
            raise RuntimeError("pywin32 未安装，无法操作 Word")

    async def word_replace(self, session_id: str, placeholder: str, value: str) -> int:
        s = self._get_session(session_id, "word")
        # 需要先激活文档
        s["doc"].Activate()
        find = s["doc"].Content.Find
        find.Text = placeholder
        find.Replacement.Text = value
        find.Forward = True
        find.Wrap = 1  # wdFindContinue
        find.Execute(Replace=2)  # wdReplaceAll
        # 无法直接获取替换次数，返回 -1 表示已执行
        return -1

    async def word_fill_table(self, session_id: str, table_index: int, data: TableData) -> int:
        s = self._get_session(session_id, "word")
        s["doc"].Activate()

        try:
            table = s["doc"].Tables(table_index)  # 1-indexed
        except Exception:
            return 0

        # 写表头
        for j, h in enumerate(data.headers):
            table.Cell(1, j + 1).Range.Text = str(h)

        # 写数据行
        for i, row in enumerate(data.rows):
            for j, val in enumerate(row):
                table.Cell(i + 2, j + 1).Range.Text = str(val)

        return len(data.rows)

    async def word_refresh_fields(self, session_id: str) -> None:
        s = self._get_session(session_id, "word")
        s["doc"].ActiveWindow.View.Type = 3  # wdPrintView
        s["doc"].Fields.Update()

    async def word_save_as(self, session_id: str, filepath: str) -> None:
        s = self._get_session(session_id, "word")
        s["doc"].SaveAs2(filepath)

    async def word_close(self, session_id: str) -> None:
        s = self._get_session(session_id, "word")
        try:
            s["doc"].Close(SaveChanges=False)
        except Exception:
            pass
        # 不退出 Word Application，因为可能还有别的文档在操作
        self._sessions.pop(session_id, None)

    # ═══════════════════════════════════════════════
    # 图表联动
    # ═══════════════════════════════════════════════

    async def excel_range_copy_to_word(
        self, excel_session: str, range_spec: str,
        word_session: str, word_bookmark: str = ""
    ) -> bool:
        """
        Excel 区域 → Word 表格。逐格写入，不依赖剪贴板。
        """
        xl_s = self._get_session(excel_session, "excel")
        wd_s = self._get_session(word_session, "word")

        # 获取 Excel 区域和数据
        ws = None
        if xl_s["mode"] == "hot":
            ws = xl_s["wb"].ActiveSheet
        if not ws:
            try:
                import win32com.client
                excel = self._com_get_active("Excel.Application", "ET.Application")
                ws = excel.ActiveWorkbook.ActiveSheet
            except Exception:
                return False

        rng = ws.Range(range_spec)
        data = rng.Value  # tuple of tuples
        rows = len(data)
        cols = len(data[0]) if rows else 1

        # 定位 Word 插入点
        wd_s["doc"].Activate()
        selection = wd_s["app"].Selection
        if word_bookmark and wd_s["doc"].Bookmarks.Exists(word_bookmark):
            wd_s["doc"].Bookmarks(word_bookmark).Select()
        else:
            wd_s["doc"].Content.Select()
            selection.Collapse(Direction=0)  # wdCollapseEnd
            selection.TypeParagraph()  # 换行

        # 创建 Word 表格
        table = wd_s["doc"].Tables.Add(
            selection.Range, rows, cols,
            DefaultTableBehavior=2  # wdWord9TableBehavior
        )

        # 逐格填入数据
        for r_i in range(rows):
            for c_i in range(cols):
                val = data[r_i][c_i] if r_i < len(data) and c_i < len(data[r_i]) else ""
                try:
                    cell = table.Cell(r_i + 1, c_i + 1)
                    cell.Range.Text = str(val) if val is not None else ""
                except Exception:
                    continue

        # 自动调整宽度
        try:
            table.AutoFitBehavior(2)
        except Exception:
            pass

        return True

    async def excel_chart_copy_to_word(
        self, excel_session: str, chart_index: int,
        word_session: str, word_bookmark: str = ""
    ) -> bool:
        """
        Excel 图表 → 复制到 Word 指定位置。
        chart_index: 图表序号（从1开始）
        word_bookmark: Word 中的书签名（可选）
        """
        xl_s = self._get_session(excel_session, "excel")
        wd_s = self._get_session(word_session, "word")

        if xl_s["mode"] != "hot":
            # 尝试获取活动 Excel
            try:
                import win32com.client
                excel = self._com_get_active("Excel.Application", "ET.Application")
                wb = excel.ActiveWorkbook
            except Exception:
                return False
        else:
            wb = xl_s["wb"]

        try:
            chart = wb.Charts(chart_index)
            chart.Copy()
        except Exception:
            # 尝试从工作表内嵌图表复制
            try:
                ws = wb.ActiveSheet
                chart_obj = ws.ChartObjects(chart_index)
                chart_obj.Copy()
            except Exception:
                return False

        wd_s["doc"].Activate()
        selection = wd_s["app"].Selection

        if word_bookmark and wd_s["doc"].Bookmarks.Exists(word_bookmark):
            wd_s["doc"].Bookmarks(word_bookmark).Select()

        selection.Paste()
        return True

    async def word_export_pdf(self, session_id: str, pdf_path: str) -> bool:
        """Word 导出为 PDF。自动启动 Word 如果没开。"""
        from pathlib import Path
        s = self._get_session(session_id, "word")
        path = Path(pdf_path)
        if path.suffix.lower() != ".pdf":
            path = path.with_suffix(".pdf")
        path.parent.mkdir(parents=True, exist_ok=True)
        try:
            # 先保存 doc
            s["doc"].Save()
            s["doc"].ExportAsFixedFormat(str(path), 17)  # wdFormatPDF
            import time
            time.sleep(0.5)
            return path.exists()
        except Exception:
            return False

    # ═══════════════════════════════════════════════
    # 工具方法
    # ═══════════════════════════════════════════════

    def _get_session(self, session_id: str, expected_type: str) -> dict:
        s = self._sessions.get(session_id)
        if not s:
            raise ValueError(f"会话不存在: {session_id}")
        if s["type"] != expected_type:
            raise ValueError(f"会话类型错误: 期望 {expected_type}, 实际 {s['type']}")
        return s

    @staticmethod
    def _parse_cell(cell: str) -> tuple[int, int]:
        """'A1' → (1, 1), 'AB12' → (28, 12)"""
        col_str = ""
        row_str = ""
        for ch in cell.upper():
            if ch.isalpha():
                col_str += ch
            else:
                row_str += ch
        col = 0
        for ch in col_str:
            col = col * 26 + (ord(ch) - ord('A') + 1)
        return col, int(row_str)

    @staticmethod
    def _col_letter(col: int) -> str:
        """1→A, 28→AB"""
        result = ""
        while col > 0:
            col, rem = divmod(col - 1, 26)
            result = chr(rem + ord('A')) + result
        return result

    @staticmethod
    def _is_file_locked_by_excel(filepath: str) -> bool:
        """检测文件是否被 Excel 锁定。"""
        try:
            # 尝试以写模式打开——如果被 Excel 锁了会报错
            f = open(filepath, "a")
            f.close()
            return False
        except (IOError, PermissionError):
            return True

    @staticmethod
    def _find_open_workbook(filepath: str) -> tuple:
        """在所有 Excel 进程中搜索已打开的文件。返回 (workbook, app) 或 (None, None)。"""
        try:
            import pythoncom
            pythoncom.CoInitialize()
            import win32com.client
            try:
                excel = self._com_get_active("Excel.Application", "ET.Application")
            except Exception:
                return None, None

            path = Path(filepath).resolve()
            for wb in excel.Workbooks:
                try:
                    if Path(wb.FullName).resolve() == path:
                        return wb, excel
                except Exception:
                    continue
            return None, None
        except ImportError:
            return None, None
